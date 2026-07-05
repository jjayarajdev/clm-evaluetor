"""Polished Evaluetor value calculator — Excel workbook for discovery calls.

Five sheets:
1. Conversation Script — how to use this in a 30-minute discovery call
2. Value Calculator     — live input/output (Conservative / Base / Upside)
3. Pricing Examples     — 3-portfolio reference table
4. Assumptions          — leakage & recovery scenario library
5. Glossary             — what each term means, defensibly

Output: sales/Evaluetor-Value-Calculator.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Brand colors (Evaluetor palette) ────────────────────────────
PAPER    = "F7F6F3"
PAPER_2  = "EFEDE8"
INK      = "0E0E10"
INK_2    = "2D2F36"
INK_3    = "5B5F66"
RULE_2   = "DDDBD5"
ORANGE   = "E94E1B"
ORANGE_2 = "C73E0F"

hairline = Side(style="thin", color=RULE_2)
heavyline = Side(style="medium", color=INK)
border_b = Border(bottom=hairline)

def set_cell(ws, ref, value, *, font=None, fill=None, align=None, number_format=None, border=None):
    c = ws[ref]
    c.value = value
    if font: c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    if number_format: c.number_format = number_format
    if border: c.border = border
    return c

def width(ws, **cols):
    for col, w in cols.items():
        ws.column_dimensions[col].width = w

def merge_set(ws, range_str, value, **kwargs):
    ws.merge_cells(range_str)
    first = range_str.split(":")[0]
    set_cell(ws, first, value, **kwargs)

def label_font(size=10, color=INK_3, bold=True):
    return Font(name="Calibri", size=size, color=color, bold=bold)

def body_font(size=11, color=INK_2, bold=False, italic=False):
    return Font(name="Calibri", size=size, color=color, bold=bold, italic=italic)

def title_font(size=22, color=INK, bold=True):
    return Font(name="Calibri", size=size, color=color, bold=bold)

def section_font(size=13, color=INK, bold=True):
    return Font(name="Calibri", size=size, color=color, bold=bold)

def fill_color(hex_color):
    return PatternFill("solid", fgColor=hex_color)

wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
wrap_left_mid = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_mid = Alignment(horizontal="right", vertical="center")

wb = Workbook()
default = wb.active
wb.remove(default)


# ═══════════════════════════════════════════════════════════════
# SHEET 1 · CONVERSATION SCRIPT
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Conversation Script")
ws.sheet_view.showGridLines = False
width(ws, A=4, B=42, C=42, D=4)

ws.row_dimensions[2].height = 32
merge_set(ws, "B2:C2", "Evaluetor — Discovery Call Script",
          font=title_font(size=22), align=wrap_left_mid)
ws.row_dimensions[3].height = 22
merge_set(ws, "B3:C3", "Use the Value Calculator (sheet 2) live in the call. This script is your scaffold.",
          font=body_font(size=11, color=INK_3, italic=True), align=wrap_left_mid)

ws.row_dimensions[5].height = 22
for col, k, v in [("B", "DURATION", "30 minutes"), ("C", "WHO YOU NEED", "Procurement + Finance lead")]:
    set_cell(ws, f"{col}5", k, font=label_font(color=ORANGE_2, size=9), align=wrap_left_mid)
    set_cell(ws, f"{col}6", v, font=body_font(size=12, bold=True), align=wrap_left_mid)

phases = [
    ("PHASE 01",
     "Anchor — establish portfolio scale",
     "0–5 min",
     [
       "What's the annual value of the supplier (or customer) contracts you want visibility on?",
       "Roughly how many vendors / counterparties does that span?",
       "Have you had a known leakage event in the last 12 months — missed credit, surprise auto-renewal, unbilled scope?",
     ],
     "Plug their ACV number into sheet 2 (Value Calculator, cell C6). Watch their face when the recoverable euro appears."),

    ("PHASE 02",
     "Calibrate — leakage assumption",
     "5–12 min",
     [
       "How much of your contract value would you say leaks post-signature today? Industry research puts it at 8–11%.",
       "What % of that do you think your current process catches?",
       "If you knew the leakage was X%, what would you spend to recover 30–40% of it?",
     ],
     "Toggle the Leakage scenario in sheet 2 (cell C7: Conservative / Base / Upside). Show all three live so they cannot argue 'best case.'"),

    ("PHASE 03",
     "Reframe — cost of inaction",
     "12–18 min",
     [
       "Over 36 months, with no changes, what is the cumulative leakage on this portfolio?",
       "What does ONE missed SLA credit on your largest vendor typically cost you?",
       "Who in the business currently owns the recovery of that money?",
     ],
     "Use the cumulative figure (recoverable x 3 years) to anchor the cost of doing nothing. This is the highest-impact moment in the call."),

    ("PHASE 04",
     "Propose — investment vs recovery",
     "18–24 min",
     [
       "Our typical year-one investment for a portfolio of your size is around €Y. That returns €X.",
       "Which would feel like the right pilot scope — 5 contracts, 20 contracts, or one business unit?",
       "Who else on your side needs to be in the room when we walk through the findings?",
     ],
     "Cite sheet 3 (Pricing Examples). Show 'Year 1 total' alongside 'recoverable value' so the ROI multiple is unmissable. Never quote price without the recoverable euro next to it."),

    ("PHASE 05",
     "Close — remove the risk",
     "24–30 min",
     [
       "If we did a 30-minute review on 5 of your live contracts next week, what would we need to bring?",
       "Who on your side decides whether we proceed after the review?",
       "What would prevent us moving forward after we show the findings?",
     ],
     "Use the closing line: 'If we don't find value, you don't proceed.' Removes 80% of buyer hesitation."),
]

row = 9
for label, header, time, qs, note in phases:
    ws.row_dimensions[row].height = 22
    set_cell(ws, f"B{row}", label, font=label_font(color=ORANGE_2, size=10), align=wrap_left_mid)
    set_cell(ws, f"C{row}", time, font=label_font(color=INK_3, size=10, bold=False),
             align=Alignment(horizontal="right", vertical="center"))
    ws.row_dimensions[row+1].height = 28
    merge_set(ws, f"B{row+1}:C{row+1}", header,
              font=section_font(size=15), align=wrap_left_mid)
    for i, q in enumerate(qs):
        r = row + 2 + i
        ws.row_dimensions[r].height = 38
        set_cell(ws, f"B{r}", f"Q{i+1}.", font=body_font(size=10, color=ORANGE_2, bold=True),
                 align=Alignment(horizontal="left", vertical="top"))
        set_cell(ws, f"C{r}", q, font=body_font(size=11, color=INK_2), align=wrap_left)
    r_note = row + 2 + len(qs)
    ws.row_dimensions[r_note].height = 42
    set_cell(ws, f"B{r_note}", "PRESENTER NOTE", font=label_font(color=INK_3, size=9),
             align=Alignment(horizontal="left", vertical="top"))
    set_cell(ws, f"C{r_note}", note,
             font=body_font(size=11, color=INK, italic=True),
             fill=fill_color(PAPER_2), align=wrap_left)
    row = r_note + 2

ws.row_dimensions[row].height = 22
merge_set(ws, f"B{row}:C{row}", "GOLDEN RULES",
          font=label_font(color=ORANGE_2, size=10), align=wrap_left_mid)
rules = [
    "Never lead with features. Lead with their portfolio value, your recoverable euro figure, and the ROI multiple.",
    "Never quote price without the recoverable euro figure in the same breath. Price-in-isolation always feels expensive.",
    "Never defend price by discounting. Trade: lower price for longer term, named reference, or faster decision.",
    "Always end by removing risk: 'If we don't find value, you don't proceed.'",
    "If the call wanders into 'AI agents' or 'features,' redirect: 'And what would that be worth to you on your portfolio?'",
]
for i, rule in enumerate(rules):
    r = row + 1 + i
    ws.row_dimensions[r].height = 28
    set_cell(ws, f"B{r}", "v", font=body_font(size=12, color=ORANGE_2, bold=True),
             align=Alignment(horizontal="left", vertical="top"))
    set_cell(ws, f"C{r}", rule, font=body_font(size=11, color=INK_2), align=wrap_left)


# ═══════════════════════════════════════════════════════════════
# SHEET 2 · VALUE CALCULATOR
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Value Calculator")
ws.sheet_view.showGridLines = False
width(ws, A=4, B=42, C=22, D=4, E=42, F=22, G=14, H=14)

YELLOW = fill_color("FFF4D9")

ws.row_dimensions[2].height = 32
merge_set(ws, "B2:C2", "Value Calculator",
          font=title_font(size=22), align=wrap_left_mid)
ws.row_dimensions[3].height = 22
merge_set(ws, "B3:C3", "Yellow cells are inputs. Everything else calculates automatically.",
          font=body_font(size=11, color=INK_3, italic=True), align=wrap_left_mid)

ws.row_dimensions[5].height = 22
merge_set(ws, "B5:C5", "1.  CORE INPUTS",
          font=label_font(color=INK, size=10), align=wrap_left_mid)

set_cell(ws, "B6", "Annual contract value under management (€)",
         font=body_font(size=12), align=wrap_left_mid, border=border_b)
set_cell(ws, "C6", 50_000_000, font=body_font(size=12, bold=True),
         fill=YELLOW, align=right_mid, number_format='#,##0', border=border_b)

set_cell(ws, "B7", "Leakage scenario (Conservative / Base / Upside)",
         font=body_font(size=12), align=wrap_left_mid, border=border_b)
set_cell(ws, "C7", "Base", font=body_font(size=12, bold=True),
         fill=YELLOW, align=right_mid, border=border_b)

set_cell(ws, "B8", "Leakage %", font=body_font(size=11, color=INK_3),
         align=wrap_left_mid, border=border_b)
ws["C8"] = "=IFERROR(VLOOKUP(C7,Assumptions!A2:C4,2,FALSE),0)"
ws["C8"].font = body_font(size=11, color=INK_3)
ws["C8"].alignment = right_mid
ws["C8"].number_format = '0.0%'
ws["C8"].border = border_b

set_cell(ws, "B9", "Recoverable share", font=body_font(size=11, color=INK_3),
         align=wrap_left_mid, border=border_b)
ws["C9"] = "=IFERROR(VLOOKUP(C7,Assumptions!A2:C4,3,FALSE),0)"
ws["C9"].font = body_font(size=11, color=INK_3)
ws["C9"].alignment = right_mid
ws["C9"].number_format = '0.0%'
ws["C9"].border = border_b

ws.row_dimensions[11].height = 22
merge_set(ws, "B11:C11", "2.  COMMERCIAL INPUTS",
          font=label_font(color=INK, size=10), align=wrap_left_mid)

inputs = [
    (12, "Annual platform fee (€)", 80_000),
    (13, "Implementation fee, year 1 only (€)", 40_000),
    (14, "Connector / integration fees, year 1 (€)", 30_000),
]
for r, label, val in inputs:
    set_cell(ws, f"B{r}", label, font=body_font(size=12), align=wrap_left_mid, border=border_b)
    set_cell(ws, f"C{r}", val, font=body_font(size=12, bold=True),
             fill=YELLOW, align=right_mid, number_format='#,##0', border=border_b)

ws.row_dimensions[16].height = 22
merge_set(ws, "B16:C16", "3.  CALCULATIONS",
          font=label_font(color=INK, size=10), align=wrap_left_mid)

calcs = [
    (17, "Estimated leakage value (€)",      "=C6*C8",       '#,##0',    False, INK),
    (18, "Estimated recoverable value (€)",  "=C17*C9",      '#,##0',    False, INK),
    (19, "Year 1 investment (€)",            "=SUM(C12:C14)",'#,##0',    False, INK),
    (20, "Net year 1 value (€)",             "=C18-C19",     '#,##0',    False, INK),
    (21, "ROI multiple (recoverable / investment)", "=IFERROR(C18/C19,0)", '0.0"x"', True, ORANGE_2),
    (22, "Cost of doing nothing — 3 years (€)", "=C17*3", '#,##0', True, ORANGE_2),
]
for r, label, formula, fmt, bold, color in calcs:
    set_cell(ws, f"B{r}", label,
             font=body_font(size=12 if bold else 11, bold=bold, color=INK_2),
             align=wrap_left_mid, border=border_b)
    ws[f"C{r}"] = formula
    ws[f"C{r}"].font = body_font(size=14 if bold else 12, bold=bold, color=color)
    ws[f"C{r}"].alignment = right_mid
    ws[f"C{r}"].number_format = fmt
    ws[f"C{r}"].border = border_b

# Right-side guide
ws.row_dimensions[5].height = 22
merge_set(ws, "E5:F5", "MEETING INTERPRETATION GUIDE",
          font=label_font(color=INK, size=10), align=wrap_left_mid)

interp = [
    ("If ROI > 10x", "Strong CFO case — proceed to scoping.",  ORANGE_2),
    ("If ROI 5–10x", "Viable. Reinforce speed-to-value and risk reduction.", INK),
    ("If ROI < 5x", "Reduce implementation scope OR expand portfolio in scope.", INK_3),
]
for i, (k, v, color) in enumerate(interp):
    r = 6 + i * 2
    set_cell(ws, f"E{r}", k, font=body_font(size=13, bold=True, color=color), align=wrap_left_mid)
    set_cell(ws, f"E{r+1}", v, font=body_font(size=11, color=INK_2, italic=True), align=wrap_left)

ws.row_dimensions[14].height = 22
merge_set(ws, "E14:F14", "RULE OF THUMB FOR THE ROOM",
          font=label_font(color=INK, size=10), align=wrap_left_mid)
merge_set(ws, "E15:F15",
          "Recoverable value ≈ 3.15% of annual contract value (base case = 9% leakage x 35% recovery).",
          font=body_font(size=12, color=INK, italic=True), align=wrap_left)
ws.row_dimensions[15].height = 38

ws.row_dimensions[18].height = 22
merge_set(ws, "E18:F18", "THE OPENING LINE TO USE",
          font=label_font(color=ORANGE_2, size=10), align=wrap_left_mid)
merge_set(ws, "E19:F22",
          "\"We typically identify €1–2M of recoverable value per €50M of contract volume. Our fee is a fraction of that.\"",
          font=body_font(size=13, color=INK, italic=True, bold=True), align=wrap_left)

ws.row_dimensions[24].height = 22
merge_set(ws, "B24:H24", "4.  THREE-CASE SENSITIVITY",
          font=label_font(color=INK, size=10), align=wrap_left_mid)

headers = ["Scenario", "Leakage %", "Recoverable %", "Recoverable value (€)", "Year 1 investment (€)", "Net value (€)", "ROI multiple"]
for i, h in enumerate(headers):
    col = get_column_letter(2 + i)
    set_cell(ws, f"{col}25", h,
             font=label_font(color=INK, size=9), align=wrap_left_mid,
             border=Border(bottom=heavyline))
ws.row_dimensions[25].height = 22

scenarios = [
    ("Conservative", 0.08, 0.30, INK_3),
    ("Base",         0.09, 0.35, ORANGE_2),
    ("Upside",       0.11, 0.40, INK),
]
for i, (name, lk, rec, color) in enumerate(scenarios):
    r = 26 + i
    set_cell(ws, f"B{r}", name, font=body_font(size=12, bold=True, color=color),
             align=wrap_left_mid, border=border_b)
    set_cell(ws, f"C{r}", lk, font=body_font(size=11, color=INK_2),
             align=right_mid, number_format='0.0%', border=border_b)
    set_cell(ws, f"D{r}", rec, font=body_font(size=11, color=INK_2),
             align=right_mid, number_format='0.0%', border=border_b)
    ws[f"E{r}"] = f"=C6*C{r}*D{r}"
    ws[f"E{r}"].font = body_font(size=12, bold=True, color=color)
    ws[f"E{r}"].alignment = right_mid
    ws[f"E{r}"].number_format = '#,##0'
    ws[f"E{r}"].border = border_b
    ws[f"F{r}"] = "=C19"
    ws[f"F{r}"].font = body_font(size=11, color=INK_2)
    ws[f"F{r}"].alignment = right_mid
    ws[f"F{r}"].number_format = '#,##0'
    ws[f"F{r}"].border = border_b
    ws[f"G{r}"] = f"=E{r}-F{r}"
    ws[f"G{r}"].font = body_font(size=11, color=INK_2)
    ws[f"G{r}"].alignment = right_mid
    ws[f"G{r}"].number_format = '#,##0'
    ws[f"G{r}"].border = border_b
    ws[f"H{r}"] = f"=IFERROR(E{r}/F{r},0)"
    ws[f"H{r}"].font = body_font(size=12, bold=True, color=color)
    ws[f"H{r}"].alignment = right_mid
    ws[f"H{r}"].number_format = '0.0"x"'
    ws[f"H{r}"].border = border_b


# ═══════════════════════════════════════════════════════════════
# SHEET 3 · PRICING EXAMPLES
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Pricing Examples")
ws.sheet_view.showGridLines = False
width(ws, A=4, B=22, C=18, D=22, E=18, F=20, G=22, H=14)

ws.row_dimensions[2].height = 32
merge_set(ws, "B2:F2", "Pricing Examples — by portfolio size",
          font=title_font(size=22), align=wrap_left_mid)
ws.row_dimensions[3].height = 22
merge_set(ws, "B3:F3", "Three reference deal shapes. Use these to anchor pricing in conversation.",
          font=body_font(size=11, color=INK_3, italic=True), align=wrap_left_mid)

headers = ["Portfolio (€)", "Platform fee (€)", "Implementation (€)", "Connectors (€)", "Year 1 total (€)", "Recoverable — base (€)", "ROI"]
for i, h in enumerate(headers):
    col = get_column_letter(2 + i)
    set_cell(ws, f"{col}5", h,
             font=label_font(color=INK, size=9), align=wrap_left_mid,
             border=Border(bottom=heavyline))
ws.row_dimensions[5].height = 22

rows = [
    (20_000_000, 60_000, 20_000, 20_000),
    (50_000_000, 80_000, 40_000, 30_000),
    (150_000_000, 220_000, 80_000, 50_000),
]
for i, (acv, pf, impl, conn) in enumerate(rows):
    r = 6 + i
    set_cell(ws, f"B{r}", acv, font=body_font(size=13, bold=True),
             align=right_mid, number_format='#,##0', border=border_b)
    set_cell(ws, f"C{r}", pf, font=body_font(size=12),
             align=right_mid, number_format='#,##0', border=border_b)
    set_cell(ws, f"D{r}", impl, font=body_font(size=12),
             align=right_mid, number_format='#,##0', border=border_b)
    set_cell(ws, f"E{r}", conn, font=body_font(size=12),
             align=right_mid, number_format='#,##0', border=border_b)
    ws[f"F{r}"] = f"=SUM(C{r}:E{r})"
    ws[f"F{r}"].font = body_font(size=13, bold=True, color=INK)
    ws[f"F{r}"].alignment = right_mid
    ws[f"F{r}"].number_format = '#,##0'
    ws[f"F{r}"].border = border_b
    ws[f"G{r}"] = f"=B{r}*0.09*0.35"
    ws[f"G{r}"].font = body_font(size=13, bold=True, color=ORANGE_2)
    ws[f"G{r}"].alignment = right_mid
    ws[f"G{r}"].number_format = '#,##0'
    ws[f"G{r}"].border = border_b
    ws[f"H{r}"] = f"=IFERROR(G{r}/F{r},0)"
    ws[f"H{r}"].font = body_font(size=14, bold=True, color=ORANGE_2)
    ws[f"H{r}"].alignment = right_mid
    ws[f"H{r}"].number_format = '0.0"x"'
    ws[f"H{r}"].border = border_b
    ws.row_dimensions[r].height = 26

# Tier table
ws.row_dimensions[11].height = 22
merge_set(ws, "B11:H11", "PRICING TIERS — WHAT YOU ACTUALLY SELL",
          font=label_font(color=INK, size=10), align=wrap_left_mid)

for i, h in enumerate(["Tier", "Range", "What's included", "Sell to"]):
    col = ["B", "C", "D", "G"][i]
    set_cell(ws, f"{col}12", h, font=label_font(color=INK_3, size=9),
             align=wrap_left_mid, border=Border(bottom=heavyline))
ws.row_dimensions[12].height = 22

tier_rows = [
    ("Tier 01 · Monitoring",   "€40–60k",     "Ingestion, extraction, risk scoring, renewal alerts, basic dashboard", "Procurement / Ops"),
    ("Tier 02 · Recovery",     "€60–120k",    "+ SLA monitoring, obligation tracking, credit calculation, breach alerts", "Procurement + Finance"),
    ("Tier 03 · Optimisation", "€120–250k+",  "+ Relationship governance, vendor scoring, forecasting, AI recommendations", "CFO / COO"),
]
for i, (tier, rng, incl, sell) in enumerate(tier_rows):
    r = 13 + i
    ws.row_dimensions[r].height = 36
    is_recovery = "Recovery" in tier
    color = ORANGE_2 if is_recovery else INK
    set_cell(ws, f"B{r}", tier, font=body_font(size=12, bold=True, color=color),
             align=wrap_left_mid, border=border_b)
    set_cell(ws, f"C{r}", rng, font=body_font(size=12, color=INK_2),
             align=wrap_left_mid, border=border_b)
    merge_set(ws, f"D{r}:F{r}", incl,
              font=body_font(size=11, color=INK_2), align=wrap_left)
    set_cell(ws, f"G{r}", sell,
             font=body_font(size=11, color=INK_2, italic=True),
             align=wrap_left_mid, border=border_b)

# Add-ons
ws.row_dimensions[17].height = 22
merge_set(ws, "B17:H17", "ADD-ONS & FLOORS",
          font=label_font(color=INK, size=10), align=wrap_left_mid)
addons = [
    ("Connector — ERP (SAP, Oracle, NetSuite)", "€20k / year"),
    ("Connector — ITSM (ServiceNow)",            "€15k / year"),
    ("Connector — Finance / FX",                  "€10k / year"),
    ("Implementation — Lite (<= 50 contracts)",   "€15k one-off"),
    ("Implementation — Standard",                 "€25–50k one-off"),
    ("Implementation — Enterprise",               "€50–120k one-off"),
    ("Platform floor — Mid-market minimum",       "€40k / year"),
    ("Platform floor — Enterprise minimum",       "€80k / year"),
]
for i, (k, v) in enumerate(addons):
    r = 18 + i
    ws.row_dimensions[r].height = 22
    merge_set(ws, f"B{r}:F{r}", k, font=body_font(size=11, color=INK_2), align=wrap_left_mid)
    merge_set(ws, f"G{r}:H{r}", v, font=body_font(size=11, bold=True, color=INK), align=right_mid)
    for c in range(2, 9):
        ws.cell(row=r, column=c).border = border_b


# ═══════════════════════════════════════════════════════════════
# SHEET 4 · ASSUMPTIONS
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Assumptions")
ws.sheet_view.showGridLines = False
width(ws, A=18, B=16, C=18, D=60)

set_cell(ws, "A1", "Scenario", font=label_font(color=INK, size=9),
         align=wrap_left_mid, border=Border(bottom=heavyline))
set_cell(ws, "B1", "Leakage %", font=label_font(color=INK, size=9),
         align=right_mid, border=Border(bottom=heavyline))
set_cell(ws, "C1", "Recoverable %", font=label_font(color=INK, size=9),
         align=right_mid, border=Border(bottom=heavyline))
set_cell(ws, "D1", "Source note", font=label_font(color=INK, size=9),
         align=wrap_left_mid, border=Border(bottom=heavyline))

scenarios = [
    ("Conservative", 0.08, 0.30, "Conservative end of industry range — used when buyer is sceptical or portfolio is well-managed already."),
    ("Base",         0.09, 0.35, "Base case — typical post-signature leakage across European mid-market portfolios with active management uplift."),
    ("Upside",       0.11, 0.40, "Upside case — applies to portfolios with high SLA spend, indexed pricing, multi-vendor complexity, no active execution programme today."),
]
for i, (name, lk, rec, note) in enumerate(scenarios):
    r = 2 + i
    ws.row_dimensions[r].height = 42
    set_cell(ws, f"A{r}", name, font=body_font(size=12, bold=True),
             align=wrap_left_mid, border=border_b)
    set_cell(ws, f"B{r}", lk, font=body_font(size=12),
             align=right_mid, number_format='0%', border=border_b)
    set_cell(ws, f"C{r}", rec, font=body_font(size=12),
             align=right_mid, number_format='0%', border=border_b)
    set_cell(ws, f"D{r}", note, font=body_font(size=11, color=INK_2, italic=True),
             align=wrap_left, border=border_b)


# ═══════════════════════════════════════════════════════════════
# SHEET 5 · GLOSSARY
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet("Glossary")
ws.sheet_view.showGridLines = False
width(ws, A=4, B=28, C=70)

ws.row_dimensions[2].height = 32
merge_set(ws, "B2:C2", "Glossary",
          font=title_font(size=22), align=wrap_left_mid)
ws.row_dimensions[3].height = 22
merge_set(ws, "B3:C3", "What each term means — defensibly. Use these definitions in conversation.",
          font=body_font(size=11, color=INK_3, italic=True), align=wrap_left_mid)

terms = [
    ("Annual contract value (ACVUM)",
     "Total annual euro value of the supplier or customer contracts in scope for visibility. Not lifetime contract value, not enterprise revenue — only the portfolio under management."),
    ("Leakage",
     "The gap between the value a contract was signed for and the value the buyer (or seller) actually realises. Sources: missed obligations, unenforced SLAs, accidental renewals, off-contract spend, scope creep, missed rebates."),
    ("Recoverable share",
     "Portion of identified leakage that can be closed under active management. 30–40% is the published industry benchmark; we use 35% in the base case."),
    ("Year 1 investment",
     "Platform fee + implementation + connector / integration fees. After year 1, only the platform fee plus connector fees recur."),
    ("ROI multiple",
     "Recoverable value / year 1 investment. CFO mental model: < 3x reject · 5–10x consider · > 10x approve."),
    ("Cost of doing nothing",
     "Leakage value x number of years the gap persists. Used to anchor the urgency of acting now vs deferring."),
    ("Tier 01 · Monitoring",
     "Entry product. Contract ingestion, extraction, risk scoring, renewal alerts, basic dashboard. The land deal."),
    ("Tier 02 · Recovery",
     "Core revenue tier. Adds SLA monitoring, obligation tracking, credit calculation, breach alerts. Where the credit math closes the ROI gap."),
    ("Tier 03 · Optimisation",
     "Strategic tier. Adds relationship governance, vendor scoring, forecasting, AI recommendations. CFO/COO level."),
    ("Connector",
     "An integration with an operational system (ServiceNow, SAP, Oracle, central bank FX feed). Priced separately because it requires scoped engineering work."),
    ("Perception gap",
     "The numerical difference between how the buyer scores a vendor relationship and how the vendor scores the same relationship — measured along the same KPI definitions. Closes the loop on relationship governance."),
]
for i, (term, defn) in enumerate(terms):
    r = 5 + i
    ws.row_dimensions[r].height = 56
    set_cell(ws, f"B{r}", term, font=body_font(size=12, bold=True),
             align=Alignment(horizontal="left", vertical="top"), border=border_b)
    set_cell(ws, f"C{r}", defn, font=body_font(size=11, color=INK_2),
             align=wrap_left, border=border_b)


wb.active = 0

import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "Evaluetor-Value-Calculator.xlsx")
wb.save(out)
print(f"Saved: {out}")
print(f"Sheets: {len(wb.sheetnames)} — {wb.sheetnames}")
