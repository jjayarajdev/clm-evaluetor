"""Specialized Excel SLA Parser for IT Outsourcing Contracts.

Parses structured SLA data from Excel files like:
- Service Level Matrix (Attachment 3-A)
- Service Levels Outcome Examples (Attachment 3-D)
"""

import logging
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, BinaryIO

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


@dataclass
class ParsedSLAMetric:
    """A single SLA metric extracted from Excel."""

    section_reference: str | None = None
    sla_name: str = ""
    sla_description: str | None = None
    category: str | None = None  # Sheet name or group
    service_tower: str | None = None  # Parent service group

    expected_value: Decimal | None = None
    minimum_value: Decimal | None = None
    measurement_window: str | None = None  # monthly, quarterly, etc.

    at_risk_percentage: Decimal | None = None
    effective_date: str | None = None  # "Eff + mos**" column

    is_critical: bool = False
    has_earnback: bool = False

    source_sheet: str = ""
    source_row: int = 0


@dataclass
class ParsedSLAExample:
    """An SLA outcome example with monthly performance data."""

    example_number: int = 0
    description: str = ""
    monthly_values: dict[str, Decimal] = field(default_factory=dict)  # {"Jan": 0.96, "Feb": 0.97, ...}
    average_value: Decimal | None = None

    has_default: bool = False
    has_earnback: bool = False
    minimum_threshold: Decimal | None = None

    source_sheet: str = ""


@dataclass
class ExcelSLAParseResult:
    """Result of parsing an SLA Excel file."""

    filename: str
    file_type: str  # "service_level_matrix", "outcome_examples", "unknown"

    metrics: list[ParsedSLAMetric] = field(default_factory=list)
    examples: list[ParsedSLAExample] = field(default_factory=list)

    total_at_risk: Decimal | None = None  # Total supplier at risk
    at_risk_pool_available: Decimal | None = None

    parse_errors: list[str] = field(default_factory=list)
    success: bool = True


class ExcelSLAParser:
    """Parser for SLA-related Excel files."""

    # Patterns to detect SLA-related files
    SLA_FILE_PATTERNS = [
        r"service.?level",
        r"sla",
        r"attachment.?3",
        r"performance.?metric",
    ]

    # Month abbreviations for parsing
    MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    def __init__(self):
        self._sla_patterns = [re.compile(p, re.IGNORECASE) for p in self.SLA_FILE_PATTERNS]

    def is_sla_file(self, filename: str) -> bool:
        """Check if filename suggests SLA-related content."""
        return any(p.search(filename) for p in self._sla_patterns)

    def detect_file_type(self, filename: str, wb: openpyxl.Workbook) -> str:
        """Detect the type of SLA file based on content."""
        filename_lower = filename.lower()

        if "matrix" in filename_lower or "3-a" in filename_lower:
            return "service_level_matrix"
        if "outcome" in filename_lower or "example" in filename_lower or "3-d" in filename_lower:
            return "outcome_examples"

        # Check sheet names
        sheet_names = [s.lower() for s in wb.sheetnames]
        if any("critical" in s or "key measurement" in s for s in sheet_names):
            return "service_level_matrix"
        if any("example" in s for s in sheet_names):
            return "outcome_examples"

        return "unknown"

    def parse_file(self, file_path: str | Path) -> ExcelSLAParseResult:
        """Parse an Excel file for SLA data."""
        path = Path(file_path)
        result = ExcelSLAParseResult(
            filename=path.name,
            file_type="unknown"
        )

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            result.file_type = self.detect_file_type(path.name, wb)

            if result.file_type == "service_level_matrix":
                self._parse_service_level_matrix(wb, result)
            elif result.file_type == "outcome_examples":
                self._parse_outcome_examples(wb, result)
            else:
                # Try both parsers
                self._parse_service_level_matrix(wb, result)
                self._parse_outcome_examples(wb, result)

            wb.close()

        except Exception as e:
            logger.exception(f"Error parsing SLA Excel file {path.name}: {e}")
            result.success = False
            result.parse_errors.append(str(e))

        return result

    def parse_bytes(self, file: BinaryIO, filename: str = "document.xlsx") -> ExcelSLAParseResult:
        """Parse Excel bytes for SLA data."""
        result = ExcelSLAParseResult(
            filename=filename,
            file_type="unknown"
        )

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            result.file_type = self.detect_file_type(filename, wb)

            if result.file_type == "service_level_matrix":
                self._parse_service_level_matrix(wb, result)
            elif result.file_type == "outcome_examples":
                self._parse_outcome_examples(wb, result)
            else:
                self._parse_service_level_matrix(wb, result)
                self._parse_outcome_examples(wb, result)

            wb.close()

        except Exception as e:
            logger.exception(f"Error parsing SLA Excel bytes {filename}: {e}")
            result.success = False
            result.parse_errors.append(str(e))

        return result

    def _parse_service_level_matrix(self, wb: openpyxl.Workbook, result: ExcelSLAParseResult) -> None:
        """Parse Service Level Matrix sheets."""
        for sheet_name in wb.sheetnames:
            sheet_lower = sheet_name.lower()

            # Skip cover pages
            if "cover" in sheet_lower or "title" in sheet_lower:
                continue

            ws = wb[sheet_name]

            # Detect category from sheet name
            category = None
            is_critical = False
            if "critical" in sheet_lower:
                category = "Critical Service Levels"
                is_critical = True
            elif "key" in sheet_lower:
                category = "Key Measurements"
            elif "deliverable" in sheet_lower:
                category = "Critical Deliverables"
            else:
                category = sheet_name

            self._parse_sla_sheet(ws, result, category, is_critical)

    def _parse_sla_sheet(
        self,
        ws: Worksheet,
        result: ExcelSLAParseResult,
        category: str,
        is_critical: bool = False
    ) -> None:
        """Parse a single sheet for SLA metrics."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return

        # Find header row (contains "Expected", "Minimum")
        header_row_idx = None
        expected_col = None
        minimum_col = None
        window_col = None
        name_col = None
        section_col = None

        for i, row in enumerate(rows[:15]):  # Check first 15 rows for header
            row_str = [str(c).lower() if c else "" for c in row]

            for j, cell in enumerate(row_str):
                if "expected" in cell:
                    expected_col = j
                    header_row_idx = i
                if "minimum" in cell:
                    minimum_col = j
                if "window" in cell or "measurement" in cell:
                    window_col = j
                if "section" in cell:
                    section_col = j

        if header_row_idx is None or expected_col is None:
            return

        # Find name column (usually 2-3 cols to left of expected)
        name_col = max(0, expected_col - 2)

        # Extract at-risk total if present
        for row in rows[:10]:
            row_str = " ".join([str(c) if c else "" for c in row])
            if "total" in row_str.lower() and "risk" in row_str.lower():
                for cell in row:
                    if isinstance(cell, (int, float)):
                        result.total_at_risk = self._to_decimal(cell)
                        break

        # Parse data rows
        current_service_tower = None

        for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
            if not any(row):
                continue

            # Get section reference
            section_ref = None
            if section_col is not None and section_col < len(row):
                section_ref = str(row[section_col]).strip() if row[section_col] else None
            elif len(row) > 1 and row[1]:
                # Section reference often in column B
                val = str(row[1]).strip()
                if re.match(r"^\d+(\.\d+)*$", val):
                    section_ref = val

            # Get service name
            sla_name = ""
            description = ""

            for col_idx in range(name_col, min(name_col + 3, len(row))):
                if col_idx < len(row) and row[col_idx]:
                    text = str(row[col_idx]).strip()
                    if text and not text.startswith(("0.", "1.", "Eff")):
                        if not sla_name:
                            sla_name = text
                        elif not description:
                            description = text

            if not sla_name:
                continue

            # Check if this is a service tower header (no expected value)
            expected_val = None
            minimum_val = None

            if expected_col < len(row) and row[expected_col]:
                expected_val = self._to_decimal(row[expected_col])
            if minimum_col is not None and minimum_col < len(row) and row[minimum_col]:
                minimum_val = self._to_decimal(row[minimum_col])

            if expected_val is None and minimum_val is None:
                # This might be a service tower header
                current_service_tower = sla_name
                continue

            # Get measurement window
            window = None
            if window_col is not None and window_col < len(row) and row[window_col]:
                window = str(row[window_col]).strip()

            metric = ParsedSLAMetric(
                section_reference=section_ref,
                sla_name=sla_name,
                sla_description=description if description else None,
                category=category,
                service_tower=current_service_tower,
                expected_value=expected_val,
                minimum_value=minimum_val,
                measurement_window=window,
                is_critical=is_critical,
                source_sheet=ws.title,
                source_row=row_idx,
            )

            result.metrics.append(metric)

    def _parse_outcome_examples(self, wb: openpyxl.Workbook, result: ExcelSLAParseResult) -> None:
        """Parse Outcome Examples sheets."""
        for sheet_name in wb.sheetnames:
            sheet_lower = sheet_name.lower()

            # Skip cover pages
            if "cover" in sheet_lower or "title" in sheet_lower:
                continue

            ws = wb[sheet_name]

            # Look for example data with monthly columns
            if "data" in sheet_lower or "example" in sheet_lower:
                self._parse_example_data_sheet(ws, result)

    def _parse_example_data_sheet(self, ws: Worksheet, result: ExcelSLAParseResult) -> None:
        """Parse example data with monthly performance values."""
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return

        # Find header row with month columns
        header_row_idx = None
        month_cols: dict[str, int] = {}

        for i, row in enumerate(rows[:10]):
            row_str = [str(c).strip() if c else "" for c in row]

            for j, cell in enumerate(row_str):
                for month in self.MONTHS:
                    if cell.lower().startswith(month.lower()):
                        month_cols[month] = j

            if len(month_cols) >= 3:  # Found month columns
                header_row_idx = i
                break

        if header_row_idx is None:
            return

        # Look for minimum threshold
        minimum_threshold = None
        for row in rows:
            row_str = " ".join([str(c) if c else "" for c in row])
            if "minimum" in row_str.lower() and "service level" in row_str.lower():
                for cell in row:
                    if isinstance(cell, (int, float)) and 0 < cell < 1:
                        minimum_threshold = self._to_decimal(cell)
                        break

        # Parse example rows
        for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
            if not any(row):
                continue

            # Get example number and description
            example_num = None
            description = ""

            for col_idx in range(min(3, len(row))):
                if row[col_idx]:
                    val = row[col_idx]
                    if isinstance(val, (int, float)) and 1 <= val <= 10:
                        example_num = int(val)
                    elif isinstance(val, str) and val.strip():
                        description = val.strip()

            if example_num is None or not description:
                continue

            # Parse monthly values
            monthly_values = {}
            for month, col_idx in month_cols.items():
                if col_idx < len(row) and row[col_idx]:
                    val = self._to_decimal(row[col_idx])
                    if val is not None:
                        monthly_values[month] = val

            if not monthly_values:
                continue

            # Detect default and earnback from description
            desc_lower = description.lower()
            # "No Default" means no default, "Default" alone or "with Default" means has default
            has_default = "default" in desc_lower and "no default" not in desc_lower
            # "No Earnback" means no earnback
            has_earnback = "earnback" in desc_lower and "no earnback" not in desc_lower

            example = ParsedSLAExample(
                example_number=example_num,
                description=description,
                monthly_values=monthly_values,
                average_value=self._calculate_average(monthly_values),
                has_default=has_default,
                has_earnback=has_earnback,
                minimum_threshold=minimum_threshold,
                source_sheet=ws.title,
            )

            result.examples.append(example)

    def _to_decimal(self, value: Any) -> Decimal | None:
        """Convert value to Decimal."""
        if value is None:
            return None

        try:
            if isinstance(value, str):
                # Remove % and common formatting
                value = value.strip().rstrip("%").replace(",", "")
                if not value or value.lower() in ("xx.xx", "n/a", "-"):
                    return None

            dec = Decimal(str(value))

            # Convert percentages > 1 to decimals if needed
            # But keep values like 0.97 as-is
            if dec > 1 and dec <= 100:
                dec = dec / 100

            return dec

        except (InvalidOperation, ValueError):
            return None

    def _calculate_average(self, values: dict[str, Decimal]) -> Decimal | None:
        """Calculate average of decimal values."""
        if not values:
            return None

        total = sum(values.values())
        return total / len(values)


def get_excel_sla_parser() -> ExcelSLAParser:
    """Get singleton Excel SLA parser instance."""
    return ExcelSLAParser()
